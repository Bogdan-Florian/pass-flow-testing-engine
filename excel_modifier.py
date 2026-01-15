"""
Excel Modifier
Modifies Excel files (.xlsx and .xls) before processing, including primary key auto-increment.
Compatible with CSVModifier interface.
"""

from pathlib import Path
from csv_modifier import increment_numeric_suffix


class ExcelModifier:
    """Modifies Excel files before processing"""

    def __init__(self, filepath, sheet=None, has_header=True, **kwargs):
        """
        Initialize Excel modifier.

        Args:
            filepath: Path to Excel file (.xlsx or .xls)
            sheet: Sheet name or index (default: first sheet)
            has_header: Whether Excel has a header row (default: True)
            **kwargs: Ignored (for compatibility with CSVModifier params)
        """
        self.filepath = Path(filepath)
        self.sheet = sheet
        self.has_header = has_header
        self._validate_file()

    def _validate_file(self):
        """Validate file exists and has correct extension"""
        if not self.filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {self.filepath}")

        ext = self.filepath.suffix.lower()
        if ext not in ('.xlsx', '.xls'):
            raise ValueError(f"Unsupported file format: {ext}. Use .xlsx or .xls")

    def _is_xlsx(self):
        """Check if file is .xlsx format"""
        return self.filepath.suffix.lower() == '.xlsx'

    def increment_primary_key(self, column=None, column_index=None):
        """
        Increment the numeric suffix of a primary key column by 1.

        Args:
            column: Column name (when has_header=True)
            column_index: Column index 0-based (alternative to column name)

        Returns:
            int: Number of rows modified

        Raises:
            ValueError: If neither column nor column_index specified,
                       or if column not found
        """
        if column is None and column_index is None:
            raise ValueError("Either 'column' or 'column_index' must be specified")

        if self._is_xlsx():
            return self._increment_xlsx(column, column_index)
        else:
            return self._increment_xls(column, column_index)

    def _increment_xlsx(self, column, column_index):
        """Increment primary key in xlsx file"""
        import openpyxl

        wb = openpyxl.load_workbook(self.filepath)

        if self.sheet is None:
            ws = wb.active
        elif isinstance(self.sheet, int):
            ws = wb.worksheets[self.sheet]
        else:
            ws = wb[self.sheet]

        # Get all rows as list for easier manipulation
        rows = list(ws.iter_rows())
        if not rows:
            wb.close()
            raise ValueError("Excel file is empty")

        # Determine target column index
        if self.has_header:
            headers = [cell.value for cell in rows[0]]
            if column is not None:
                if column not in headers:
                    wb.close()
                    raise ValueError(f"Column '{column}' not found in Excel headers: {headers}")
                target_index = headers.index(column)
            else:
                target_index = column_index
                if target_index >= len(headers):
                    wb.close()
                    raise ValueError(f"Column index {column_index} out of range (max: {len(headers)-1})")
            data_start = 1
        else:
            target_index = column_index if column_index is not None else 0
            if target_index >= len(rows[0]):
                wb.close()
                raise ValueError(f"Column index {target_index} out of range (max: {len(rows[0])-1})")
            data_start = 0

        # Increment values
        errors = []
        modified = 0
        for row_idx in range(data_start, len(rows)):
            row = rows[row_idx]
            if target_index >= len(row):
                continue

            cell = row[target_index]
            original_value = cell.value

            if original_value is None or str(original_value).strip() == '':
                continue

            try:
                new_value = increment_numeric_suffix(str(original_value))
                # Write to actual cell in worksheet
                ws.cell(row=row_idx + 1, column=target_index + 1, value=new_value)
                modified += 1
            except ValueError as e:
                row_num = row_idx + 1
                errors.append(f"Row {row_num}: {e}")

        if errors:
            for error in errors:
                print(f"  WARNING: {error}")

        wb.save(self.filepath)
        wb.close()

        return modified

    def _increment_xls(self, column, column_index):
        """
        Increment primary key in xls file.
        Note: xlrd is read-only, so we need to use xlwt or convert to xlsx.
        For simplicity, we'll convert to xlsx, modify, then save back as xls.
        """
        import xlrd

        # Read the xls file
        rb = xlrd.open_workbook(self.filepath, formatting_info=True)

        if self.sheet is None:
            rs = rb.sheet_by_index(0)
            sheet_idx = 0
        elif isinstance(self.sheet, int):
            rs = rb.sheet_by_index(self.sheet)
            sheet_idx = self.sheet
        else:
            rs = rb.sheet_by_name(self.sheet)
            sheet_idx = rb.sheet_names().index(self.sheet)

        if rs.nrows == 0:
            raise ValueError("Excel file is empty")

        # Determine target column
        if self.has_header:
            headers = [rs.cell_value(0, c) for c in range(rs.ncols)]
            if column is not None:
                if column not in headers:
                    raise ValueError(f"Column '{column}' not found in Excel headers: {headers}")
                target_index = headers.index(column)
            else:
                target_index = column_index
                if target_index >= len(headers):
                    raise ValueError(f"Column index {column_index} out of range (max: {len(headers)-1})")
            data_start = 1
        else:
            target_index = column_index if column_index is not None else 0
            if target_index >= rs.ncols:
                raise ValueError(f"Column index {target_index} out of range (max: {rs.ncols-1})")
            data_start = 0

        # Read all data and prepare modifications
        try:
            import xlwt
            from xlutils.copy import copy as xlcopy
        except ImportError:
            # If xlwt not available, convert to xlsx format
            print("  NOTE: Converting .xls to .xlsx for modification (xlwt not installed)")
            return self._convert_and_increment_xls(column, column_index)

        # Copy workbook for writing
        wb = xlcopy(rb)
        ws = wb.get_sheet(sheet_idx)

        errors = []
        modified = 0
        for row_idx in range(data_start, rs.nrows):
            original_value = rs.cell_value(row_idx, target_index)

            if original_value is None or str(original_value).strip() == '':
                continue

            try:
                new_value = increment_numeric_suffix(str(original_value))
                ws.write(row_idx, target_index, new_value)
                modified += 1
            except ValueError as e:
                row_num = row_idx + 1
                errors.append(f"Row {row_num}: {e}")

        if errors:
            for error in errors:
                print(f"  WARNING: {error}")

        wb.save(str(self.filepath))

        return modified

    def _convert_and_increment_xls(self, column, column_index):
        """
        Fallback: Convert xls to xlsx, modify, convert back.
        Used when xlwt/xlutils not available.
        """
        import xlrd
        import openpyxl

        # Read xls
        rb = xlrd.open_workbook(self.filepath)

        if self.sheet is None:
            rs = rb.sheet_by_index(0)
            sheet_name = rb.sheet_names()[0]
        elif isinstance(self.sheet, int):
            rs = rb.sheet_by_index(self.sheet)
            sheet_name = rb.sheet_names()[self.sheet]
        else:
            rs = rb.sheet_by_name(self.sheet)
            sheet_name = self.sheet

        # Create new xlsx workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Copy data
        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                value = rs.cell_value(row_idx, col_idx)
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

        # Save as xlsx temporarily
        xlsx_path = self.filepath.with_suffix('.xlsx')
        wb.save(xlsx_path)
        wb.close()

        # Modify using xlsx modifier
        self.filepath = xlsx_path
        modified = self._increment_xlsx(column, column_index)

        # Convert back to xls (keep as xlsx since xlwt not available)
        print(f"  NOTE: File saved as {xlsx_path} (xlwt not installed for .xls write)")

        return modified
