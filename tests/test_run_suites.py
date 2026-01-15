import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

import run_suites


def write_text(path, content):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def create_xlsx(filepath, data):
    """Helper to create test xlsx file"""
    import openpyxl
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(filepath)
    wb.close()
    return filepath


def create_xlsx_multi_sheet(filepath, sheets_data):
    """Helper to create xlsx with multiple sheets. sheets_data is dict {sheet_name: data}"""
    import openpyxl
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    for i, (sheet_name, data) in enumerate(sheets_data.items()):
        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        for row_idx, row in enumerate(data, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)
    wb.close()
    return filepath


def yaml_path(path):
    return f"'{Path(path).as_posix()}'"


class FakeValidator:
    fail_ids = set()

    def __init__(self, *args, **kwargs):
        pass

    def validate_row(self, row_num, row_data, variables_config, validations):
        policy_id = row_data.get("id")
        if policy_id in self.fail_ids:
            return {
                "has_failures": True,
                "validations": [{
                    "name": "check",
                    "passed": False,
                    "errors": ["forced failure"],
                    "sql_executed": "select 1"
                }]
            }
        return {
            "has_failures": False,
            "validations": [{
                "name": "check",
                "passed": True,
                "errors": [],
                "sql_executed": "select 1"
            }]
        }


class FakeBatchExecutor:
    def __init__(self, *args, **kwargs):
        pass

    def execute_batches(self, *args, **kwargs):
        return True, []


def build_minimal_config(csv_rel_path):
    return f"""file:
  path: {csv_rel_path}
  delimiter: ","
  encoding: "utf-8"

validations:
  - name: check
    sql: "select 1"
    expect:
      row_count: 1
"""


def build_manifest(config_path, report_dir, db_url="sqlite:///:memory:"):
    return f"""version: 1

database:
  connection_url: '{db_url}'

suites:
  - name: suite_one
    enabled: true
    critical: false
    config: {yaml_path(config_path)}

reporting:
  output_dir: {yaml_path(report_dir)}
"""


def test_run_suites_cli_smoke(tmp_path, monkeypatch):
    manifest_path = tmp_path / "manifest.yaml"
    write_text(
        manifest_path,
        "version: 1\n"
        "database:\n"
        "  connection_url: 'sqlite:///:memory:'\n"
        "suites:\n"
        "  - name: s\n"
        "    config: 'c.yaml'\n"
    )

    def fake_run_all(self, suite_filter=None, tags_filter=None):
        return True

    monkeypatch.setattr(run_suites.SuiteRunner, "run_all", fake_run_all)
    monkeypatch.setattr(sys, "argv", ["run_suites.py", str(manifest_path)])

    with pytest.raises(SystemExit) as excinfo:
        run_suites.main()

    assert excinfo.value.code == 0


def test_path_resolution_and_report_output(tmp_path, monkeypatch):
    (tmp_path / "other").mkdir(parents=True, exist_ok=True)
    suite_dir = tmp_path / "suites" / "nested"
    csv_path = suite_dir / "data.csv"
    write_text(csv_path, "id\n1\n")

    config_path = suite_dir / "config.yaml"
    write_text(config_path, build_minimal_config("data.csv"))

    report_dir = tmp_path / "reports"
    manifest_path = tmp_path / "manifest.yaml"
    write_text(manifest_path, build_manifest(config_path, report_dir))

    FakeValidator.fail_ids = set()
    monkeypatch.setattr(run_suites, "Validator", FakeValidator)
    monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)
    monkeypatch.chdir(tmp_path / "other")

    runner = run_suites.SuiteRunner(manifest_path)
    assert runner.run_all()

    report_file = report_dir / "suite_one" / "suite_one_results.json"
    report = report_file.read_text(encoding="utf-8")
    assert "\"total_rows\": 1" in report


def test_report_includes_sql_executed_on_failure(tmp_path, monkeypatch):
    suite_dir = tmp_path / "suite"
    csv_path = suite_dir / "data.csv"
    write_text(csv_path, "id\nFAIL\n")

    config_path = suite_dir / "config.yaml"
    write_text(config_path, build_minimal_config("data.csv"))

    report_dir = tmp_path / "reports"
    manifest_path = tmp_path / "manifest.yaml"
    write_text(manifest_path, build_manifest(config_path, report_dir))

    FakeValidator.fail_ids = {"FAIL"}
    monkeypatch.setattr(run_suites, "Validator", FakeValidator)
    monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

    runner = run_suites.SuiteRunner(manifest_path)
    assert not runner.run_all()

    report_file = report_dir / "suite_one" / "suite_one_results.json"
    report = report_file.read_text(encoding="utf-8")
    assert "\"sql_executed\": \"select 1\"" in report


def test_aggregate_summary_two_suites(tmp_path, monkeypatch):
    report_dir = tmp_path / "reports"

    suite_a = tmp_path / "suite_a"
    write_text(suite_a / "data.csv", "id\nPASS\n")
    write_text(suite_a / "config.yaml", build_minimal_config("data.csv"))

    suite_b = tmp_path / "suite_b"
    write_text(suite_b / "data.csv", "id\nFAIL\n")
    write_text(suite_b / "config.yaml", build_minimal_config("data.csv"))

    manifest_path = tmp_path / "manifest.yaml"
    write_text(
        manifest_path,
        f"""version: 1

database:
  connection_url: 'sqlite:///:memory:'

suites:
  - name: suite_a
    enabled: true
    critical: false
    config: {yaml_path(suite_a / 'config.yaml')}
  - name: suite_b
    enabled: true
    critical: false
    config: {yaml_path(suite_b / 'config.yaml')}

reporting:
  output_dir: {yaml_path(report_dir)}
"""
    )

    FakeValidator.fail_ids = {"FAIL"}
    monkeypatch.setattr(run_suites, "Validator", FakeValidator)
    monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

    runner = run_suites.SuiteRunner(manifest_path)
    assert not runner.run_all()

    aggregate = (report_dir / "aggregate_summary.json").read_text(encoding="utf-8")
    assert "\"total_suites\": 2" in aggregate
    assert "\"passed_suites\": 1" in aggregate
    assert "\"failed_suites\": 1" in aggregate


def test_batch_failure_short_circuit(tmp_path, monkeypatch):
    class FailingBatchExecutor:
        def __init__(self, *args, **kwargs):
            pass

        def execute_batches(self, *args, **kwargs):
            return False, [{
                "batch_name": "step1",
                "script": "script",
                "status": "FAILED",
                "exit_code": 1,
                "error": "boom",
                "duration_seconds": 0.1
            }]

    suite_dir = tmp_path / "suite"
    write_text(suite_dir / "data.csv", "id\n1\n")
    write_text(suite_dir / "config.yaml", build_minimal_config("data.csv") + "batches:\n  - name: step1\n    script: step1\n")

    report_dir = tmp_path / "reports"
    manifest_path = tmp_path / "manifest.yaml"
    write_text(manifest_path, build_manifest(suite_dir / "config.yaml", report_dir))

    monkeypatch.setattr(run_suites, "Validator", FakeValidator)
    monkeypatch.setattr(run_suites, "BatchExecutor", FailingBatchExecutor)

    runner = run_suites.SuiteRunner(manifest_path)
    assert not runner.run_all()

    report_file = report_dir / "suite_one" / "suite_one_results.json"
    report = report_file.read_text(encoding="utf-8")
    assert "\"batch_execution\"" in report
    assert "\"total_rows\": 0" in report


# =============================================================================
# EXCEL INTEGRATION TESTS
# =============================================================================

def build_excel_config(xlsx_rel_path, sheet=None):
    """Build config for Excel file"""
    sheet_line = f"  sheet: \"{sheet}\"\n" if sheet else ""
    return f"""file:
  path: {xlsx_rel_path}
  has_header: true
{sheet_line}
validations:
  - name: check
    sql: "select 1"
    expect:
      row_count: 1
"""


def build_excel_config_no_header(xlsx_rel_path):
    """Build config for Excel file without headers"""
    return f"""file:
  path: {xlsx_rel_path}
  has_header: false

validations:
  - name: check
    sql: "select 1"
    expect:
      row_count: 1
"""


class TestExcelIntegration:
    """Integration tests for Excel file support"""

    def test_basic_excel_xlsx_file(self, tmp_path, monkeypatch):
        """Test that basic Excel .xlsx file works end-to-end"""
        suite_dir = tmp_path / "suite"
        xlsx_path = suite_dir / "data.xlsx"
        create_xlsx(xlsx_path, [
            ["id", "name", "amount"],
            ["1", "Alice", "100"],
            ["2", "Bob", "200"],
        ])

        config_path = suite_dir / "config.yaml"
        write_text(config_path, build_excel_config("data.xlsx"))

        report_dir = tmp_path / "reports"
        manifest_path = tmp_path / "manifest.yaml"
        write_text(manifest_path, build_manifest(config_path, report_dir))

        FakeValidator.fail_ids = set()
        monkeypatch.setattr(run_suites, "Validator", FakeValidator)
        monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

        runner = run_suites.SuiteRunner(manifest_path)
        assert runner.run_all()

        report_file = report_dir / "suite_one" / "suite_one_results.json"
        report = report_file.read_text(encoding="utf-8")
        assert "\"total_rows\": 2" in report

    def test_excel_with_specific_sheet(self, tmp_path, monkeypatch):
        """Test that specifying a sheet name works correctly"""
        suite_dir = tmp_path / "suite"
        xlsx_path = suite_dir / "data.xlsx"
        create_xlsx_multi_sheet(xlsx_path, {
            "Sheet1": [["wrong"], ["data"]],
            "TestData": [["id", "value"], ["1", "test"], ["2", "data"], ["3", "more"]],
        })

        config_path = suite_dir / "config.yaml"
        write_text(config_path, build_excel_config("data.xlsx", sheet="TestData"))

        report_dir = tmp_path / "reports"
        manifest_path = tmp_path / "manifest.yaml"
        write_text(manifest_path, build_manifest(config_path, report_dir))

        FakeValidator.fail_ids = set()
        monkeypatch.setattr(run_suites, "Validator", FakeValidator)
        monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

        runner = run_suites.SuiteRunner(manifest_path)
        assert runner.run_all()

        report_file = report_dir / "suite_one" / "suite_one_results.json"
        report = report_file.read_text(encoding="utf-8")
        # Should have 3 data rows from TestData sheet
        assert "\"total_rows\": 3" in report

    def test_excel_without_headers(self, tmp_path, monkeypatch):
        """Test Excel file with has_header: false"""
        suite_dir = tmp_path / "suite"
        xlsx_path = suite_dir / "data.xlsx"
        # No header row - all data
        create_xlsx(xlsx_path, [
            ["VAL1", "100", "Active"],
            ["VAL2", "200", "Pending"],
        ])

        config_path = suite_dir / "config.yaml"
        write_text(config_path, build_excel_config_no_header("data.xlsx"))

        report_dir = tmp_path / "reports"
        manifest_path = tmp_path / "manifest.yaml"
        write_text(manifest_path, build_manifest(config_path, report_dir))

        FakeValidator.fail_ids = set()
        monkeypatch.setattr(run_suites, "Validator", FakeValidator)
        monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

        runner = run_suites.SuiteRunner(manifest_path)
        assert runner.run_all()

        report_file = report_dir / "suite_one" / "suite_one_results.json"
        report = report_file.read_text(encoding="utf-8")
        # Should have 2 rows (no header to skip)
        assert "\"total_rows\": 2" in report

    def test_excel_validation_failure(self, tmp_path, monkeypatch):
        """Test that Excel validation failures are reported correctly"""
        suite_dir = tmp_path / "suite"
        xlsx_path = suite_dir / "data.xlsx"
        create_xlsx(xlsx_path, [
            ["id", "status"],
            ["PASS", "ok"],
            ["FAIL", "bad"],
        ])

        config_path = suite_dir / "config.yaml"
        write_text(config_path, build_excel_config("data.xlsx"))

        report_dir = tmp_path / "reports"
        manifest_path = tmp_path / "manifest.yaml"
        write_text(manifest_path, build_manifest(config_path, report_dir))

        FakeValidator.fail_ids = {"FAIL"}
        monkeypatch.setattr(run_suites, "Validator", FakeValidator)
        monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

        runner = run_suites.SuiteRunner(manifest_path)
        assert not runner.run_all()  # Should fail

        report_file = report_dir / "suite_one" / "suite_one_results.json"
        report = report_file.read_text(encoding="utf-8")
        assert "\"passed_rows\": 1" in report
        assert "\"failed_rows\": 1" in report

    def test_mixed_csv_and_excel_suites(self, tmp_path, monkeypatch):
        """Test running suites with both CSV and Excel files"""
        report_dir = tmp_path / "reports"

        # Suite A uses CSV
        suite_a = tmp_path / "suite_a"
        write_text(suite_a / "data.csv", "id\nCSV1\nCSV2\n")
        write_text(suite_a / "config.yaml", build_minimal_config("data.csv"))

        # Suite B uses Excel
        suite_b = tmp_path / "suite_b"
        create_xlsx(suite_b / "data.xlsx", [
            ["id"],
            ["XLS1"],
            ["XLS2"],
            ["XLS3"],
        ])
        write_text(suite_b / "config.yaml", build_excel_config("data.xlsx"))

        manifest_path = tmp_path / "manifest.yaml"
        write_text(
            manifest_path,
            f"""version: 1

database:
  connection_url: 'sqlite:///:memory:'

suites:
  - name: csv_suite
    enabled: true
    critical: false
    config: {yaml_path(suite_a / 'config.yaml')}
  - name: excel_suite
    enabled: true
    critical: false
    config: {yaml_path(suite_b / 'config.yaml')}

reporting:
  output_dir: {yaml_path(report_dir)}
"""
        )

        FakeValidator.fail_ids = set()
        monkeypatch.setattr(run_suites, "Validator", FakeValidator)
        monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

        runner = run_suites.SuiteRunner(manifest_path)
        assert runner.run_all()

        # Check CSV suite report
        csv_report = (report_dir / "csv_suite" / "csv_suite_results.json").read_text(encoding="utf-8")
        assert "\"total_rows\": 2" in csv_report

        # Check Excel suite report
        excel_report = (report_dir / "excel_suite" / "excel_suite_results.json").read_text(encoding="utf-8")
        assert "\"total_rows\": 3" in excel_report

        # Check aggregate
        aggregate = (report_dir / "aggregate_summary.json").read_text(encoding="utf-8")
        assert "\"total_suites\": 2" in aggregate
        assert "\"passed_suites\": 2" in aggregate

    def test_excel_with_auto_increment(self, tmp_path, monkeypatch):
        """Test that auto-increment works with Excel files"""
        suite_dir = tmp_path / "suite"
        xlsx_path = suite_dir / "data.xlsx"
        create_xlsx(xlsx_path, [
            ["OrderID", "Amount"],
            ["ORD-0001", "100"],
            ["ORD-0002", "200"],
        ])

        config_path = suite_dir / "config.yaml"
        write_text(config_path, f"""file:
  path: data.xlsx
  has_header: true

primary_key:
  column: OrderID
  auto_increment: true

validations:
  - name: check
    sql: "select 1"
    expect:
      row_count: 1
""")

        report_dir = tmp_path / "reports"
        manifest_path = tmp_path / "manifest.yaml"
        write_text(manifest_path, build_manifest(config_path, report_dir))

        FakeValidator.fail_ids = set()
        monkeypatch.setattr(run_suites, "Validator", FakeValidator)
        monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

        runner = run_suites.SuiteRunner(manifest_path)
        assert runner.run_all()

        # Verify the Excel file was modified (auto-incremented)
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        assert ws.cell(2, 1).value == "ORD-0002"  # Was ORD-0001
        assert ws.cell(3, 1).value == "ORD-0003"  # Was ORD-0002
        wb.close()

    def test_csv_with_sheet_param_in_config_is_ignored(self, tmp_path, monkeypatch):
        """Test that sheet parameter in config doesn't break CSV processing"""
        suite_dir = tmp_path / "suite"
        csv_path = suite_dir / "data.csv"
        write_text(csv_path, "id\n1\n2\n")

        # Config has sheet parameter but file is CSV - should be ignored
        config_path = suite_dir / "config.yaml"
        write_text(config_path, f"""file:
  path: data.csv
  delimiter: ","
  has_header: true
  sheet: "SomeSheet"

validations:
  - name: check
    sql: "select 1"
    expect:
      row_count: 1
""")

        report_dir = tmp_path / "reports"
        manifest_path = tmp_path / "manifest.yaml"
        write_text(manifest_path, build_manifest(config_path, report_dir))

        FakeValidator.fail_ids = set()
        monkeypatch.setattr(run_suites, "Validator", FakeValidator)
        monkeypatch.setattr(run_suites, "BatchExecutor", FakeBatchExecutor)

        runner = run_suites.SuiteRunner(manifest_path)
        assert runner.run_all()  # Should not crash

        report_file = report_dir / "suite_one" / "suite_one_results.json"
        report = report_file.read_text(encoding="utf-8")
        assert "\"total_rows\": 2" in report
