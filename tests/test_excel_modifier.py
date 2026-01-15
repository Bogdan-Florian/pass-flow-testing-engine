"""
Tests for ExcelModifier - Primary Key Auto-Increment for Excel files
"""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_modifier import ExcelModifier


def create_xlsx(tmp_path, filename, data):
    """Helper to create test xlsx file"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    filepath = tmp_path / filename
    wb.save(filepath)
    wb.close()
    return filepath


def read_xlsx(filepath):
    """Helper to read xlsx file contents"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    data = []
    for row in ws.iter_rows():
        data.append([cell.value for cell in row])
    wb.close()
    return data


class TestExcelModifierWithHeaders:
    """Tests for ExcelModifier with header row"""

    def test_increment_by_column_name(self, tmp_path):
        data = [
            ["OrderID", "Amount", "Status"],
            ["ORD-0001", 100, "ACTIVE"],
            ["ORD-0002", 200, "PENDING"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        rows = modifier.increment_primary_key(column="OrderID")

        assert rows == 2
        result = read_xlsx(filepath)
        assert result[1][0] == "ORD-0002"
        assert result[2][0] == "ORD-0003"

    def test_increment_by_column_index(self, tmp_path):
        data = [
            ["ID", "Name", "Value"],
            ["REC001", "Test", 100],
            ["REC002", "Demo", 200],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        rows = modifier.increment_primary_key(column_index=0)

        assert rows == 2
        result = read_xlsx(filepath)
        assert result[1][0] == "REC002"
        assert result[2][0] == "REC003"

    def test_column_not_found_raises(self, tmp_path):
        data = [
            ["ID", "Name"],
            ["001", "Test"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        with pytest.raises(ValueError, match="Column 'OrderID' not found"):
            modifier.increment_primary_key(column="OrderID")

    def test_column_index_out_of_range_raises(self, tmp_path):
        data = [
            ["ID", "Name"],
            ["001", "Test"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        with pytest.raises(ValueError, match="out of range"):
            modifier.increment_primary_key(column_index=5)

    def test_neither_column_nor_index_raises(self, tmp_path):
        data = [["ID"], ["001"]]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        with pytest.raises(ValueError, match="Either 'column' or 'column_index'"):
            modifier.increment_primary_key()


class TestExcelModifierWithoutHeaders:
    """Tests for ExcelModifier without header row"""

    def test_increment_by_index_no_header(self, tmp_path):
        data = [
            ["POL-0001", 100, "ACTIVE"],
            ["POL-0002", 200, "PENDING"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=False)
        rows = modifier.increment_primary_key(column_index=0)

        assert rows == 2
        result = read_xlsx(filepath)
        assert result[0][0] == "POL-0002"
        assert result[1][0] == "POL-0003"

    def test_increment_second_column_no_header(self, tmp_path):
        data = [
            ["Name", "ID001", "Value"],
            ["Test", "ID002", 100],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=False)
        modifier.increment_primary_key(column_index=1)

        result = read_xlsx(filepath)
        assert result[0][1] == "ID002"
        assert result[1][1] == "ID003"


class TestExcelModifierEdgeCases:
    """Edge case tests"""

    def test_empty_xlsx_raises(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        filepath = tmp_path / "empty.xlsx"
        wb.save(filepath)
        wb.close()

        modifier = ExcelModifier(filepath, has_header=True)
        with pytest.raises(ValueError, match="empty"):
            modifier.increment_primary_key(column_index=0)

    def test_header_only_xlsx(self, tmp_path):
        data = [["ID", "Name", "Value"]]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        rows = modifier.increment_primary_key(column="ID")

        assert rows == 0  # No data rows to increment

    def test_empty_value_skipped(self, tmp_path):
        data = [
            ["ID", "Name"],
            ["KEY001", "Test"],
            [None, "Empty"],
            ["KEY003", "Data"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        rows = modifier.increment_primary_key(column="ID")

        result = read_xlsx(filepath)
        assert result[1][0] == "KEY002"
        assert result[2][0] is None  # Empty value preserved
        assert result[3][0] == "KEY004"

    def test_non_numeric_suffix_warns_but_continues(self, tmp_path, capsys):
        data = [
            ["ID", "Name"],
            ["KEY001", "Test"],
            ["ABC-XYZ", "NoNum"],  # No numeric suffix
            ["KEY003", "Data"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        rows = modifier.increment_primary_key(column="ID")

        assert rows == 2  # Only 2 were successfully incremented
        captured = capsys.readouterr()
        assert "WARNING" in captured.out
        assert "No numeric suffix" in captured.out

        result = read_xlsx(filepath)
        assert result[1][0] == "KEY002"
        assert result[2][0] == "ABC-XYZ"  # Unchanged
        assert result[3][0] == "KEY004"

    def test_preserves_other_data(self, tmp_path):
        data = [
            ["ID", "Name", "Amount"],
            ["KEY0001", "Alice", 100.50],
            ["KEY0002", "Bob", 200.75],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)
        modifier.increment_primary_key(column="ID")

        result = read_xlsx(filepath)
        # Check header preserved
        assert result[0] == ["ID", "Name", "Amount"]
        # Check other columns preserved
        assert result[1][1] == "Alice"
        assert result[1][2] == 100.50
        assert result[2][1] == "Bob"
        assert result[2][2] == 200.75

    def test_multiple_runs_increment_further(self, tmp_path):
        data = [
            ["ID", "Value"],
            ["KEY0001", 100],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = ExcelModifier(filepath, has_header=True)

        # First increment
        modifier.increment_primary_key(column="ID")
        result1 = read_xlsx(filepath)
        assert result1[1][0] == "KEY0002"

        # Second increment
        modifier.increment_primary_key(column="ID")
        result2 = read_xlsx(filepath)
        assert result2[1][0] == "KEY0003"

        # Third increment
        modifier.increment_primary_key(column="ID")
        result3 = read_xlsx(filepath)
        assert result3[1][0] == "KEY0004"

    def test_specific_sheet(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.cell(1, 1, "ID")
        ws1.cell(2, 1, "A001")

        ws2 = wb.create_sheet("Data")
        ws2.cell(1, 1, "OrderID")
        ws2.cell(2, 1, "ORD-0001")

        filepath = tmp_path / "multi_sheet.xlsx"
        wb.save(filepath)
        wb.close()

        # Modify specific sheet
        modifier = ExcelModifier(filepath, sheet="Data", has_header=True)
        modifier.increment_primary_key(column="OrderID")

        # Verify
        wb = openpyxl.load_workbook(filepath)
        # Sheet1 should be unchanged
        assert wb["Sheet1"].cell(2, 1).value == "A001"
        # Data sheet should be incremented
        assert wb["Data"].cell(2, 1).value == "ORD-0002"
        wb.close()


class TestFileProcessorFactory:
    """Test the file processor factory"""

    def test_get_processor_xlsx(self, tmp_path):
        from file_processor import get_processor
        data = [["Name"], ["Alice"]]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        processor = get_processor(filepath, has_header=True)
        assert processor.__class__.__name__ == "ExcelProcessor"

    def test_get_processor_csv(self, tmp_path):
        from file_processor import get_processor
        filepath = tmp_path / "test.csv"
        filepath.write_text("Name\nAlice\n", encoding="utf-8")

        processor = get_processor(filepath, has_header=True)
        assert processor.__class__.__name__ == "CSVProcessor"

    def test_get_modifier_xlsx(self, tmp_path):
        from file_processor import get_modifier
        data = [["ID"], ["001"]]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        modifier = get_modifier(filepath, has_header=True)
        assert modifier.__class__.__name__ == "ExcelModifier"

    def test_get_modifier_csv(self, tmp_path):
        from file_processor import get_modifier
        filepath = tmp_path / "test.csv"
        filepath.write_text("ID\n001\n", encoding="utf-8")

        modifier = get_modifier(filepath, has_header=True)
        assert modifier.__class__.__name__ == "CSVModifier"

    def test_unsupported_format_raises(self, tmp_path):
        from file_processor import get_processor
        filepath = tmp_path / "test.json"
        filepath.write_text("{}", encoding="utf-8")

        with pytest.raises(ValueError, match="Unsupported file format"):
            get_processor(filepath)
