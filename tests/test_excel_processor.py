"""
Tests for ExcelProcessor - supports .xlsx and .xls files
"""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_processor import ExcelProcessor


def create_xlsx(tmp_path, filename, data, has_header=True):
    """Helper to create test xlsx file"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    filepath = tmp_path / filename
    wb.save(filepath)
    wb.close()
    return filepath


class TestExcelProcessorXlsx:
    """Tests for .xlsx files"""

    def test_count_rows_with_header(self, tmp_path):
        data = [
            ["Name", "Age", "City"],
            ["Alice", 30, "NYC"],
            ["Bob", 25, "LA"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=True)
        assert processor.count_rows() == 2

    def test_count_rows_without_header(self, tmp_path):
        data = [
            ["Alice", 30, "NYC"],
            ["Bob", 25, "LA"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=False)
        assert processor.count_rows() == 2

    def test_read_rows_with_header(self, tmp_path):
        data = [
            ["Name", "Age", "City"],
            ["Alice", 30, "NYC"],
            ["Bob", 25, "LA"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=True)
        rows = list(processor.read_rows())

        assert len(rows) == 2
        assert rows[0] == (1, {"Name": "Alice", "Age": "30", "City": "NYC"})
        assert rows[1] == (2, {"Name": "Bob", "Age": "25", "City": "LA"})

    def test_read_rows_without_header(self, tmp_path):
        data = [
            ["Alice", 30, "NYC"],
            ["Bob", 25, "LA"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=False)
        rows = list(processor.read_rows())

        assert len(rows) == 2
        assert rows[0] == (1, {"0": "Alice", "1": "30", "2": "NYC"})
        assert rows[1] == (2, {"0": "Bob", "1": "25", "2": "LA"})

    def test_get_headers_with_header(self, tmp_path):
        data = [
            ["Name", "Age", "City"],
            ["Alice", 30, "NYC"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=True)
        headers = processor.get_headers()

        assert headers == ["Name", "Age", "City"]

    def test_get_headers_without_header(self, tmp_path):
        data = [
            ["Alice", 30, "NYC"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=False)
        headers = processor.get_headers()

        assert headers == ["0", "1", "2"]

    def test_specific_sheet_by_name(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.cell(1, 1, "A")
        ws1.cell(2, 1, "B")

        ws2 = wb.create_sheet("Data")
        ws2.cell(1, 1, "Name")
        ws2.cell(2, 1, "Charlie")

        filepath = tmp_path / "multi_sheet.xlsx"
        wb.save(filepath)
        wb.close()

        processor = ExcelProcessor(filepath, sheet="Data", has_header=True)
        rows = list(processor.read_rows())

        assert len(rows) == 1
        assert rows[0] == (1, {"Name": "Charlie"})

    def test_specific_sheet_by_index(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.cell(1, 1, "A")

        ws2 = wb.create_sheet("Sheet2")
        ws2.cell(1, 1, "Name")
        ws2.cell(2, 1, "Dave")

        filepath = tmp_path / "multi_sheet.xlsx"
        wb.save(filepath)
        wb.close()

        processor = ExcelProcessor(filepath, sheet=1, has_header=True)
        rows = list(processor.read_rows())

        assert len(rows) == 1
        assert rows[0] == (1, {"Name": "Dave"})

    def test_numeric_values_preserved(self, tmp_path):
        data = [
            ["ID", "Amount", "Qty"],
            ["ORD-001", 150.50, 5],
            ["ORD-002", 99.99, 10],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=True)
        rows = list(processor.read_rows())

        assert rows[0][1]["Amount"] == "150.5"
        assert rows[0][1]["Qty"] == "5"
        assert rows[1][1]["Amount"] == "99.99"

    def test_empty_cells_handled(self, tmp_path):
        data = [
            ["A", "B", "C"],
            ["1", None, "3"],
            [None, "2", None],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=True)
        rows = list(processor.read_rows())

        assert rows[0] == (1, {"A": "1", "B": "", "C": "3"})
        assert rows[1] == (2, {"A": "", "B": "2", "C": ""})


class TestExcelProcessorCompatibility:
    """Test CSVProcessor compatibility"""

    def test_ignores_csv_params(self, tmp_path):
        """Test that CSV-specific params are ignored without error"""
        data = [["Name"], ["Alice"]]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)

        # These params should be ignored for Excel
        processor = ExcelProcessor(
            filepath,
            delimiter="|",  # Ignored
            encoding="utf-8",  # Ignored
            has_header=True
        )
        rows = list(processor.read_rows())
        assert len(rows) == 1

    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            ExcelProcessor(tmp_path / "nonexistent.xlsx")

    def test_unsupported_format(self, tmp_path):
        filepath = tmp_path / "test.txt"
        filepath.write_text("data")

        with pytest.raises(ValueError, match="Unsupported file format"):
            ExcelProcessor(filepath)


class TestExcelProcessorEdgeCases:
    """Edge case tests"""

    def test_empty_file(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Don't write any data
        filepath = tmp_path / "empty.xlsx"
        wb.save(filepath)
        wb.close()

        processor = ExcelProcessor(filepath, has_header=True)
        assert processor.count_rows() == 0
        assert list(processor.read_rows()) == []

    def test_header_only(self, tmp_path):
        data = [["Name", "Age", "City"]]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=True)

        assert processor.count_rows() == 0
        assert list(processor.read_rows()) == []
        assert processor.get_headers() == ["Name", "Age", "City"]

    def test_single_row_no_header(self, tmp_path):
        data = [["Alice", 30, "NYC"]]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=False)
        rows = list(processor.read_rows())

        assert len(rows) == 1
        assert rows[0] == (1, {"0": "Alice", "1": "30", "2": "NYC"})

    def test_special_characters(self, tmp_path):
        data = [
            ["Name", "City"],
            ["Café", "München"],
            ["Tokyo", "東京"],
        ]
        filepath = create_xlsx(tmp_path, "test.xlsx", data)
        processor = ExcelProcessor(filepath, has_header=True)
        rows = list(processor.read_rows())

        assert rows[0][1]["Name"] == "Café"
        assert rows[0][1]["City"] == "München"
        assert rows[1][1]["City"] == "東京"
