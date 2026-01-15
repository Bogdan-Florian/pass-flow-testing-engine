"""
Excel Processor
Reads Excel files (.xlsx and .xls) and provides row-by-row access.
Compatible with CSVProcessor interface.
"""

from pathlib import Path


class ExcelProcessor:
    """Handles reading Excel files with or without headers"""

    def __init__(self, filepath, sheet=None, has_header=True, **kwargs):
        """
        Initialize Excel processor

        Args:
            filepath: Path to Excel file (.xlsx or .xls)
            sheet: Sheet name or index (default: first sheet)
            has_header: Whether Excel has a header row (default: True)
            **kwargs: Ignored (for compatibility with CSVProcessor params like delimiter)
        """
        self.filepath = Path(filepath)
        self.sheet = sheet
        self.has_header = has_header
        self._validate_file()

    def _validate_file(self):
        """Validate file exists and has correct extension"""
        if not self.filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {self.filepath}")

        ext = self.filepath.suffix.lower()
        if ext not in ('.xlsx', '.xls'):
            raise ValueError(f"Unsupported file format: {ext}. Use .xlsx or .xls")

    def _is_xlsx(self):
        """Check if file is .xlsx format"""
        return self.filepath.suffix.lower() == '.xlsx'

    def _open_workbook(self):
        """Open workbook and return (workbook, sheet) based on file type"""
        if self._is_xlsx():
            import openpyxl
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            if self.sheet is None:
                ws = wb.active
            elif isinstance(self.sheet, int):
                ws = wb.worksheets[self.sheet]
            else:
                ws = wb[self.sheet]
            return wb, ws, 'xlsx'
        else:
            import xlrd
            wb = xlrd.open_workbook(self.filepath)
            if self.sheet is None:
                ws = wb.sheet_by_index(0)
            elif isinstance(self.sheet, int):
                ws = wb.sheet_by_index(self.sheet)
            else:
                ws = wb.sheet_by_name(self.sheet)
            return wb, ws, 'xls'

    def count_rows(self):
        """
        Count total number of data rows (excluding header if present)

        Returns:
            int: Number of data rows
        """
        wb, ws, fmt = self._open_workbook()
        try:
            if fmt == 'xlsx':
                # For xlsx, count non-empty rows
                total = sum(1 for row in ws.iter_rows() if any(cell.value is not None for cell in row))
            else:
                # For xls
                total = ws.nrows

            if self.has_header and total > 0:
                return total - 1
            return total
        finally:
            if fmt == 'xlsx':
                wb.close()

    def read_rows(self):
        """
        Generator that yields each Excel row as a dictionary

        Yields:
            tuple: (row_number, row_data_dict)

        When has_header=True:
            row_data keys are column names from header row

        When has_header=False:
            row_data keys are column indices as strings (e.g., {'0': 'value', '1': 'value2'})
        """
        wb, ws, fmt = self._open_workbook()
        try:
            if fmt == 'xlsx':
                yield from self._read_xlsx_rows(ws)
            else:
                yield from self._read_xls_rows(ws)
        finally:
            if fmt == 'xlsx':
                wb.close()

    def _read_xlsx_rows(self, ws):
        """Read rows from xlsx worksheet"""
        rows = list(ws.iter_rows())
        if not rows:
            return

        if self.has_header:
            # First row is header
            headers = [self._cell_value(cell) for cell in rows[0]]
            headers = [str(h).strip() if h is not None else f"Column{i}" for i, h in enumerate(headers)]

            for row_number, row in enumerate(rows[1:], start=1):
                values = [self._cell_value(cell) for cell in row]
                # Skip completely empty rows
                if all(v is None or v == '' for v in values):
                    continue
                row_data = {
                    headers[i]: self._to_string(values[i])
                    for i in range(min(len(headers), len(values)))
                }
                yield row_number, row_data
        else:
            for row_number, row in enumerate(rows, start=1):
                values = [self._cell_value(cell) for cell in row]
                if all(v is None or v == '' for v in values):
                    continue
                row_data = {str(i): self._to_string(v) for i, v in enumerate(values)}
                yield row_number, row_data

    def _read_xls_rows(self, ws):
        """Read rows from xls worksheet"""
        if ws.nrows == 0:
            return

        if self.has_header:
            headers = [self._to_string(ws.cell_value(0, c)) for c in range(ws.ncols)]
            headers = [h.strip() if h else f"Column{i}" for i, h in enumerate(headers)]

            row_number = 0
            for r in range(1, ws.nrows):
                values = [ws.cell_value(r, c) for c in range(ws.ncols)]
                if all(v is None or v == '' for v in values):
                    continue
                row_number += 1
                row_data = {
                    headers[i]: self._to_string(values[i])
                    for i in range(min(len(headers), len(values)))
                }
                yield row_number, row_data
        else:
            row_number = 0
            for r in range(ws.nrows):
                values = [ws.cell_value(r, c) for c in range(ws.ncols)]
                if all(v is None or v == '' for v in values):
                    continue
                row_number += 1
                row_data = {str(i): self._to_string(v) for i, v in enumerate(values)}
                yield row_number, row_data

    def _cell_value(self, cell):
        """Extract value from openpyxl cell"""
        return cell.value

    def _to_string(self, value):
        """Convert cell value to string, handling None and numbers"""
        if value is None:
            return ''
        if isinstance(value, float):
            # Avoid scientific notation and trailing zeros
            if value == int(value):
                return str(int(value))
            return str(value)
        return str(value).strip()

    def get_headers(self):
        """
        Get list of column headers from Excel

        Returns:
            list: Column names (when has_header=True)
                  Column indices as strings (when has_header=False)
        """
        wb, ws, fmt = self._open_workbook()
        try:
            if fmt == 'xlsx':
                rows = list(ws.iter_rows(max_row=1))
                if not rows:
                    return []
                if self.has_header:
                    return [str(self._cell_value(cell)).strip() if self._cell_value(cell) else f"Column{i}"
                            for i, cell in enumerate(rows[0])]
                else:
                    return [str(i) for i in range(len(rows[0]))]
            else:
                if ws.nrows == 0:
                    return []
                if self.has_header:
                    return [str(ws.cell_value(0, c)).strip() if ws.cell_value(0, c) else f"Column{c}"
                            for c in range(ws.ncols)]
                else:
                    return [str(i) for i in range(ws.ncols)]
        finally:
            if fmt == 'xlsx':
                wb.close()
